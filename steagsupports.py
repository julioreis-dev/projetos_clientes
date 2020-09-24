import os
import pandas as pd
import time
from openpyxl import load_workbook
from datetime import datetime


def option():
    print('################################################')
    print('Digite o período de extração dos dados:')
    date1 = input('Data inicial: ')
    start = formatdates(date1)
    date2 = input('Data final: ')
    end = formatdates(date2)
    sheet_name = start + '_' + end
    return sheet_name


def option1():
    flag = True
    while flag:
        opt = [0, 1, 2, 3]
        plant = int(input('################################################'
                          '\nEscolha uma das plantas disponíveis nesta aplicação:'
                          '\nDigite 1 --> São Pedro'
                          '\nDigite 2 --> Juazeiro'
                          '\nDigite 3 --> Sol do Futuro'
                          '\nDigite 0 --> Sair.'
                          '\n################################################'
                          '\nPrezado usuário, escolha uma opção?'))
        if plant in opt:
            return plant
        else:
            print('Opção incorreta, tente novamente.')
            time.sleep(3)


def option2():
    flag = True
    while flag:
        opt2 = [0, 1, 2, 3, 4]
        plant2 = int(input('################################################'
                           '\nSelecione um tipo de equipamento disponível:'
                           '\nDigite 1 --> Inversores'
                           '\nDigite 2 --> Strings Box'
                           '\nDigite 3 --> Strings'
                           '\nDigite 4 --> Weather Station'
                           '\nDigite 0 --> Sair.'
                           '\n################################################'
                           '\nPrezado usuário, escolha uma opção?'))
        if plant2 in opt2:
            return plant2
        else:
            print('Opção incorreta, tente novamente.')
            time.sleep(3)


def readframe(df):
    df['Date'] = pd.to_datetime(df['Date'])
    df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')
    df['Time'] = pd.to_datetime(df['Time'])
    df['Time'] = df['Time'].dt.strftime('%H:%M:%S')
    df['Date'] = df['Date'] + ' ' + df['Time']
    return df


def readframe1(df):
    df['timestamp'] = pd.to_datetime(df['timestamp'])
    df['timestamp'] = df['timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S')
    df = df[['timestamp', 'Current', 'Power', 'COMS STATUS']]
    return df


def createsheets(pasta):
    if not os.path.isdir(pasta):
        os.mkdir(pasta)


def createsubsheets(pasta):
    sheet1 = os.path.join(pasta, 'são pedro')
    sheet2 = os.path.join(pasta, 'juazeiro')
    sheet3 = os.path.join(pasta, 'sol do futuro')
    sheets = [sheet1, sheet2, sheet3]
    for n in sheets:
        if not os.path.isdir(n):
            os.mkdir(n)


def organizefiles(files):
    df_data = pd.read_excel(files)
    colun = df_data.columns.values
    return colun


'Analize data of equipment inversores'


def organizetuplainverter(listcol):
    flag = True
    listdata = []
    lencollumn = len(listcol)
    index1 = 2
    index2 = 3
    while flag:
        collumntupla = (listcol[index1], listcol[index2])
        listdata.append(collumntupla)
        index1 += 2
        index2 += 2
        if index2 > lencollumn:
            flag = False
    return listdata


def organizetuplastringsbox(listcol):
    flag = True
    listdata = []
    lencollumn = len(listcol)
    index1 = 3
    index2 = 4
    index3 = 5
    while flag:
        collumntupla = (listcol[index1], listcol[index2], listcol[index3])
        listdata.append(collumntupla)
        index1 += 3
        index2 += 3
        index3 += 3
        if index2 > lencollumn:
            flag = False
    return listdata


def formatdates(date):
    timedata = datetime.strptime(date, '%d/%m/%Y').date()
    formatdata = timedata.strftime('%Y-%m-%d')
    return formatdata


def sheetperiod(*args):
    if args[0] == 1:
        sheetdest = f'{args[1]}\são pedro'
        sheets = os.path.join(sheetdest, args[2])
        if not os.path.isdir(sheets):
            os.mkdir(sheets)
    elif args[0] == 2:
        sheetdest = f'{args[1]}\juazeiro'
        sheets = os.path.join(sheetdest, args[2])
        if not os.path.isdir(sheets):
            os.mkdir(sheets)
    elif args[0] == 3:
        sheetdest = f'{args[1]}\sol do futuro'
        sheets = os.path.join(sheetdest, args[2])
        if not os.path.isdir(sheets):
            os.mkdir(sheets)


def sheetdestination(*args):
    if args[2] == 1:
        sheetdest1 = f'{args[0]}\são pedro\{args[1]}'
        return sheetdest1
    elif args[2] == 2:
        sheetdest2 = f'{args[0]}\juazeiro\{args[1]}'
        return sheetdest2
    elif args[2] == 3:
        sheetdest3 = f'{args[0]}\sol do futuro\{args[1]}'
        return sheetdest3


def stamp(*args):
    if args[2] == 1:
        namestamp1 = f'Sao Pedro-data-{args[0]}-{args[3]}-{args[1][12:21]}'
        return namestamp1
    elif args[2] == 2:
        namestamp2 = f'Juazeiro-data-{args[0]}-{args[3]}-{args[1][10:22]}'
        return namestamp2
    elif args[2] == 3:
        namestamp3 = f'Sol do Futuro-data-{args[0]}-{args[3]}-{args[1][16:27]}'
        return namestamp3


def stamp1(*args):
    if args[2] == 1:
        namestamp1 = f'Sao Pedro-data-{args[0]}-{args[3]}-{args[1][12:26]}'
        return namestamp1
    elif args[2] == 2:
        namestamp2 = f'Juazeiro-data-{args[0]}-{args[3]}-{args[1][10:22]}'
        return namestamp2
    elif args[2] == 3:
        namestamp3 = f'Sol do Futuro-data-{args[0]}-{args[3]}-{args[1][16:27]}'
        return namestamp3


def executiontime(*args):
    execution = args[0] - args[1]
    hr = execution//3600
    min = execution//60
    seg = round((execution % 60)//1, 2)
    return hr, min, seg


def workdata(*args):
    wb = load_workbook(filename=args[0])
    ws = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(ws[0])
    contline = sheet.max_row
    for line in range(2, contline+1):
        datas = sheet.cell(row=line, column=1).value

        print(type(datas))
        print(datas)
    pass